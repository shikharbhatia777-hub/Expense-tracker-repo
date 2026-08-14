from fastmcp import FastMCP
import asyncio
import os
import asyncpg
import aiofiles
import jwt
import hashlib
import secrets
import queue
import threading
import json
from datetime import datetime, timedelta, timezone as tz
from flask import Flask, request, jsonify
from flask_cors import CORS
from werkzeug.serving import run_simple
import logging

try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail
except ImportError:
    SendGridAPIClient = None
    Mail = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None

if load_dotenv is not None:
    load_dotenv()

POSTGRES_HOST = os.getenv("POSTGRES_HOST", "localhost")
POSTGRES_PORT = int(os.getenv("POSTGRES_PORT", "5432"))
POSTGRES_USER = os.getenv("POSTGRES_USER", "postgres")
POSTGRES_PASSWORD = os.getenv("POSTGRES_PASSWORD", "")
POSTGRES_DB = os.getenv("POSTGRES_DB", "postgres")

SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY", "")
SMTP_FROM = os.getenv("SMTP_FROM", "noreply@expensetracker.com")

JWT_SECRET = os.getenv("JWT_SECRET", secrets.token_urlsafe(32))
JWT_EXPIRY_HOURS = int(os.getenv("JWT_EXPIRY_HOURS", "24"))

mcp = FastMCP("ExpenseTracker")
db_lock = asyncio.Lock()
_db_initialized = False
_current_user_id = None
_db_pool = None
_email_queue = queue.Queue()
_email_worker_task = None

# Tool registry for Flask endpoints
_tool_registry = {}

# Wrap mcp.tool to also register in our registry
_original_tool = mcp.tool

def _wrapped_tool(*args, **kwargs):
    """Wrapper around mcp.tool that also registers in _tool_registry"""
    def decorator(func):
        # First register with FastMCP
        wrapped = _original_tool(*args, **kwargs)(func)
        # Then register in our registry using the function name
        _tool_registry[func.__name__] = func
        return wrapped
    return decorator

mcp.tool = _wrapped_tool


def _hash_password(password: str) -> str:
    salt = secrets.token_hex(32)
    password_hash = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
    return f"{salt}${password_hash.hex()}"


def _verify_password(password: str, hashed: str) -> bool:
    try:
        salt, password_hash = hashed.split('$')
        computed_hash = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
        return computed_hash.hex() == password_hash
    except:
        return False


def _create_token(user_id: int) -> str:
    now = datetime.now(tz.utc)
    payload = {
        'user_id': user_id,
        'exp': now + timedelta(hours=JWT_EXPIRY_HOURS),
        'iat': now
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')


def _verify_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


async def _get_connection():
    global _db_pool
    if _db_pool is None:
        raise RuntimeError("Database pool not initialized")
    return _db_pool

async def _run_with_connection(operation):
    await _ensure_db_initialized()
    conn = await _get_connection()
    return await operation(conn)


async def _execute_fetchall(conn, query, params=()):
    return await conn.fetch(query, *params)


async def _execute_fetchone(conn, query, params=()):
    return await conn.fetchrow(query, *params)


async def init_db():
    global _db_pool

    try:
        _db_pool = await asyncpg.create_pool(
            host=POSTGRES_HOST,
            port=POSTGRES_PORT,
            user=POSTGRES_USER,
            password=POSTGRES_PASSWORD,
            database=POSTGRES_DB,
            min_size=5,
            max_size=20,
            command_timeout=10,
            statement_cache_size=0,
        )

        conn = await _db_pool.acquire()
        try:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS users(
                    id SERIAL PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    email TEXT UNIQUE,
                    password_hash TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS expenses(
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    date TEXT NOT NULL,
                    amount REAL NOT NULL,
                    category TEXT NOT NULL,
                    subcategory TEXT DEFAULT '',
                    note TEXT DEFAULT '',
                    FOREIGN KEY(user_id) REFERENCES users(id)
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS credits(
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    date TEXT NOT NULL,
                    amount REAL NOT NULL,
                    source TEXT NOT NULL,
                    note TEXT DEFAULT '',
                    FOREIGN KEY(user_id) REFERENCES users(id)
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS shared_expenses(
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    date TEXT NOT NULL,
                    description TEXT,
                    total_amount REAL NOT NULL,
                    paid_by TEXT NOT NULL,
                    FOREIGN KEY(user_id) REFERENCES users(id)
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS shared_expense_participants(
                    id SERIAL PRIMARY KEY,
                    expense_id INTEGER NOT NULL,
                    participant TEXT NOT NULL,
                    share_amount REAL NOT NULL,
                    FOREIGN KEY(expense_id)
                    REFERENCES shared_expenses(id)
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS friends(
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    email TEXT,
                    FOREIGN KEY(user_id) REFERENCES users(id),
                    UNIQUE(user_id, name)
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS settlements(
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    person TEXT NOT NULL,
                    amount REAL NOT NULL,
                    settlement_date TEXT NOT NULL,
                    note TEXT DEFAULT '',
                    FOREIGN KEY(user_id) REFERENCES users(id)
                )
            """)
        finally:
            await _db_pool.release(conn)

    except Exception as e:
        raise RuntimeError(f"Unable to initialize database: {e}")


async def _normalize_payer_name(paid_by: str):
    if not paid_by:
        return "You"
    lowered = str(paid_by).strip().lower()
    if lowered in {"me", "i", "myself", "my", "self"}:
        return "You"
    return str(paid_by).strip()


def _normalize_payer_name_sync(paid_by: str):
    return asyncio.run(_normalize_payer_name(paid_by))


async def _calculate_balances(conn, user_id: int):
    balances = {}

    expense_rows = await _execute_fetchall(
        conn, "SELECT id, paid_by FROM shared_expenses WHERE user_id=$1", (user_id,)
    )

    for expense_row in expense_rows:
        expense_id, paid_by = expense_row['id'], expense_row['paid_by']
        participant_rows = await _execute_fetchall(
            conn, "SELECT participant, share_amount FROM shared_expense_participants WHERE expense_id=$1",
            (expense_id,)
        )

        if not participant_rows:
            continue

        payer_name = await _normalize_payer_name(paid_by)
        others_total = 0.0

        for participant_row in participant_rows:
            participant, share_amount = participant_row['participant'], participant_row['share_amount']
            participant_name = await _normalize_payer_name(participant)
            share_value = round(float(share_amount), 2)

            if participant_name.lower() == payer_name.lower():
                continue

            balances[participant_name] = balances.get(participant_name, 0.0) - share_value
            others_total += share_value

        if payer_name:
            balances[payer_name] = balances.get(payer_name, 0.0) + round(others_total, 2)

    # Get user's name from username for participant lookup
    user_row = await _execute_fetchone(conn, "SELECT username FROM users WHERE id=$1", (user_id,))
    username = user_row['username'] if user_row else None

    # Include expenses where this user is a participant
    if username:
        participant_expenses = await _execute_fetchall(
            conn, """
            SELECT se.id, se.paid_by, sep.share_amount
            FROM shared_expense_participants sep
            JOIN shared_expenses se ON sep.expense_id = se.id
            WHERE LOWER(sep.participant) = LOWER($1)
            """, (username,)
        )

        for expense_row in participant_expenses:
            paid_by, share_amount = expense_row['paid_by'], expense_row['share_amount']
            payer_name = await _normalize_payer_name(paid_by)
            share_value = round(float(share_amount), 2)

            if username.lower() != payer_name.lower():
                balances[payer_name] = balances.get(payer_name, 0.0) + share_value

    settlement_rows = await _execute_fetchall(
        conn, "SELECT person, amount FROM settlements WHERE user_id=$1", (user_id,)
    )

    for settlement_row in settlement_rows:
        person, amount = settlement_row['person'], settlement_row['amount']
        person_name = await _normalize_payer_name(person)
        amount_value = round(float(amount), 2)
        current_balance = balances.get(person_name, 0.0)

        # When a person settles with you, move balance towards zero
        # If they owe you (negative balance) and they pay, ADD the settlement
        # Example: -₹4,900 + ₹1,000 = -₹3,900 (still owes, but less)
        balances[person_name] = round(current_balance + amount_value, 2)

    return {name: round(balance, 2) for name, balance in balances.items()}


async def _coerce_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.endswith("%"):
            return float(text[:-1])
        return float(text)
    return None


async def _build_participant_splits(amount: float, participants: list, paid_by: str):
    if not participants:
        return []

    normalized = []
    for entry in participants:
        if isinstance(entry, str):
            normalized.append({"name": entry, "percent": None, "parts": None, "share": None})
        elif isinstance(entry, dict):
            normalized.append({
                "name": entry.get("name") or entry.get("participant") or entry.get("person"),
                "percent": await _coerce_number(entry.get("percent") or entry.get("percentage")),
                "parts": await _coerce_number(entry.get("parts") or entry.get("ratio") or entry.get("weight")),
                "share": await _coerce_number(entry.get("share") or entry.get("amount") or entry.get("value") or entry.get("owed") or entry.get("pay_amount"))
            })
        else:
            normalized.append({"name": str(entry), "percent": None, "parts": None, "share": None})

    if paid_by in [item["name"] for item in normalized]:
        total_people = len(normalized)
    else:
        total_people = len(normalized) + 1

    def _round_money(value: float):
        return round(value + 1e-9, 2)

    equal_share = _round_money(amount / total_people)

    result = []
    for item in normalized:
        name = item.get("name")
        if not name:
            continue
        percent = item.get("percent")
        parts = item.get("parts")
        share_amount = item.get("share")
        if share_amount is not None:
            share = _round_money(float(share_amount))
        elif percent is not None:
            share = _round_money(amount * float(percent) / 100.0)
        elif parts is not None:
            total_parts = sum(float(x.get("parts") or 0) for x in normalized if isinstance(x, dict) and x.get("parts") is not None)
            if total_parts <= 0:
                share = equal_share
            else:
                share = _round_money(amount * float(parts) / total_parts)
        else:
            share = equal_share
        result.append({"name": name, "share": share})

    return result


async def _build_email_summary(paid_by: str, amount: float, description: str, participant_splits: list, balances: dict, recipient_name: str):
    payer_name = await _normalize_payer_name(paid_by)
    recipient_display = recipient_name or "there"
    recipient_share = None

    for entry in participant_splits:
        if entry["name"].lower() == recipient_display.lower():
            recipient_share = entry["share"]
            break

    recipient_balance = None
    if balances is not None:
        for person, balance in balances.items():
            if str(person).lower() == recipient_display.lower():
                recipient_balance = balance
                break

    lines = []
    lines.append(f"Hi {recipient_display},")
    lines.append("")
    lines.append(f"{payer_name} paid ₹{amount:.2f} for:")
    lines.append("")
    lines.append(description)
    lines.append("")
    lines.append(f"Your share for this expense: ₹{recipient_share:.2f}" if recipient_share is not None else "Your share for this expense: ₹0.00")
    lines.append("")
    lines.append("Shared expense summary:")
    for entry in participant_splits:
        lines.append(f"- {entry['name']}: ₹{entry['share']:.2f}")
    lines.append("")
    lines.append("Please settle whenever convenient.")
    lines.append("")
    lines.append("Thank you.")
    return "\n".join(lines)


async def send_email(recipient, subject, body):
    if not recipient:
        print("[EMAIL] ❌ Skipped: no recipient provided")
        return False

    if not SENDGRID_API_KEY:
        print("[EMAIL] ❌ Skipped: SENDGRID_API_KEY not configured")
        return False

    if not SendGridAPIClient or not Mail:
        print("[EMAIL] ❌ Skipped: SendGrid library not installed")
        return False

    try:
        print(f"[EMAIL] 📧 Preparing email to {recipient}...")
        message = Mail(
            from_email=SMTP_FROM,
            to_emails=recipient,
            subject=subject,
            plain_text_content=body
        )
        print(f"[EMAIL] 🔑 Using API key: {SENDGRID_API_KEY[:20]}...")
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        print(f"[EMAIL] ✅ Email sent to {recipient} (Status: {response.status_code})")
        return True
    except Exception as e:
        print(f"[EMAIL] ❌ Failed: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False


def queue_email(recipient, subject, body):
    try:
        print(f"[QUEUE] 📬 Queuing email to {recipient}")
        print(f"[QUEUE] 📝 Subject: {subject}")
        _email_queue.put((recipient, subject, body), block=False)
        print(f"[QUEUE] ✅ Email queued successfully (Queue size now: {_email_queue.qsize()})")
    except queue.Full:
        print(f"[QUEUE] ❌ Queue full, dropping email to {recipient}")
    except Exception as e:
        print(f"[QUEUE] ❌ Error queuing: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()


def run_email_worker():
    print("[WORKER] 🚀 Email worker started")
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    print(f"[WORKER] 📊 Queue size at start: {_email_queue.qsize()}")
    while True:
        try:
            print(f"[WORKER] ⏳ Waiting for email (queue size: {_email_queue.qsize()})...")
            recipient, subject, body = _email_queue.get(timeout=1)
            print(f"[WORKER] 📨 Got email from queue! Processing to {recipient}")
            loop.run_until_complete(send_email(recipient, subject, body))
            print(f"[WORKER] ✅ Email processed successfully")
        except queue.Empty:
            pass
        except Exception as e:
            print(f"[WORKER] ❌ Critical Error: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()


async def _ensure_db_initialized():
    global _db_initialized, _email_worker_task
    if not _db_initialized:
        await init_db()
        _db_initialized = True

    # Start email worker if not already running
    if _email_worker_task is None:
        print("[INIT] 🚀 Starting email worker thread...")
        _email_worker_task = threading.Thread(target=run_email_worker, daemon=True)
        _email_worker_task.start()
        print("[INIT] ✅ Email worker thread created")


@mcp.tool(description="Register a new user with username and password.")
async def register_user(username: str, password: str, email: str = ""):
    async with db_lock:
        async def _op(conn):
            existing = await _execute_fetchone(conn, "SELECT id FROM users WHERE username=$1", (username,))
            if existing:
                return {"status": "error", "message": "Username already exists"}

            password_hash = _hash_password(password)
            created_at = datetime.now(tz.utc).isoformat()

            user_id = await conn.fetchval(
                "INSERT INTO users(username, email, password_hash, created_at) VALUES ($1,$2,$3,$4) RETURNING id",
                username, email, password_hash, created_at
            )
            return {"status": "ok", "user_id": user_id, "message": "User registered successfully"}

        return await _run_with_connection(_op)


@mcp.tool(description="Login with username and password to get a JWT token.")
async def login(username: str, password: str):
    async def _op(conn):
        user = await _execute_fetchone(conn, "SELECT id, password_hash FROM users WHERE username=$1", (username,))
        if not user:
            return {"status": "error", "message": "Invalid username or password"}

        user_id, password_hash = user['id'], user['password_hash']
        if not _verify_password(password, password_hash):
            return {"status": "error", "message": "Invalid username or password"}

        token = _create_token(user_id)
        return {"status": "ok", "token": token, "user_id": user_id, "message": "Login successful"}

    return await _run_with_connection(_op)


@mcp.tool(description="Verify a JWT token and return user information.")
async def verify_token(token: str):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        user = await _execute_fetchone(conn, "SELECT id, username, email FROM users WHERE id=$1", (payload['user_id'],))
        if not user:
            return {"status": "error", "message": "User not found"}

        return {"status": "ok", "user_id": user['id'], "username": user['username'], "email": user['email']}

    return await _run_with_connection(_op)


@mcp.tool(description="Add a new regular expense entry to the database.")
async def add_expense(token: str, date, amount, category, subcategory="", note=""):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            expense_id = await conn.fetchval(
                "INSERT INTO expenses(user_id, date, amount, category, subcategory, note) VALUES ($1,$2,$3,$4,$5,$6) RETURNING id",
                payload['user_id'], date, amount, category, subcategory, note
            )
            return {"status": "ok", "id": expense_id}

        return await _run_with_connection(_op)


@mcp.tool(description="List expense entries within an inclusive date range.")
async def list_expenses(token: str, start_date, end_date):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        rows = await _execute_fetchall(
            conn,
            """
            SELECT id, date, amount, category, subcategory, note
            FROM expenses
            WHERE user_id=$1 AND date BETWEEN $2 AND $3
            ORDER BY id ASC
            """,
            (payload['user_id'], start_date, end_date)
        )
        return [dict(r) for r in rows]

    return await _run_with_connection(_op)


@mcp.tool(description="Summarize expenses by category within an inclusive date range.")
async def summarize(token: str, start_date, end_date, category=None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        if category:
            rows = await _execute_fetchall(
                conn,
                """
                SELECT category, SUM(amount) AS total_amount
                FROM expenses
                WHERE user_id=$1 AND date BETWEEN $2 AND $3 AND category = $4
                GROUP BY category ORDER BY category ASC
                """,
                (payload['user_id'], start_date, end_date, category)
            )
        else:
            rows = await _execute_fetchall(
                conn,
                """
                SELECT category, SUM(amount) AS total_amount
                FROM expenses
                WHERE user_id=$1 AND date BETWEEN $2 AND $3
                GROUP BY category ORDER BY category ASC
                """,
                (payload['user_id'], start_date, end_date)
            )
        return [dict(r) for r in rows]

    return await _run_with_connection(_op)


@mcp.tool(description="Delete an expense using its known details instead of an ID.")
async def delete_expense(token: str, date: str, amount: float, category: str, subcategory: str = None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        if subcategory:
            rows = await _execute_fetchall(
                conn,
                """
                SELECT id, date, amount, category, subcategory, note
                FROM expenses
                WHERE user_id=$1 AND date=$2 AND amount=$3 AND category=$4 AND subcategory=$5
                """,
                (payload['user_id'], date, amount, category, subcategory)
            )
        else:
            rows = await _execute_fetchall(
                conn,
                """
                SELECT id, date, amount, category, subcategory, note
                FROM expenses
                WHERE user_id=$1 AND date=$2 AND amount=$3 AND category=$4
                """,
                (payload['user_id'], date, amount, category)
            )

        if len(rows) == 0:
            return {"status": "error", "message": "No matching expense found"}

        if len(rows) > 1:
            return {
                "status": "multiple_matches",
                "matches": [
                    {
                        "id": r['id'],
                        "date": r['date'],
                        "amount": r['amount'],
                        "category": r['category'],
                        "subcategory": r['subcategory'],
                        "note": r['note']
                    }
                    for r in rows
                ]
            }

        expense_id = rows[0]['id']
        await conn.execute("DELETE FROM expenses WHERE id=$1", expense_id)
        return {"status": "ok", "deleted_id": expense_id}

    return await _run_with_connection(_op)


@mcp.tool(description="Edit an existing expense using known details.")
async def edit_expense(token: str, old_date: str, old_amount: float, old_category: str, new_date: str = None, new_amount: float = None, new_category: str = None, new_subcategory: str = None, new_note: str = None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        rows = await _execute_fetchall(
            conn,
            """
            SELECT id, date, amount, category, subcategory, note
            FROM expenses
            WHERE user_id=$1 AND date=$2 AND amount=$3 AND category=$4
            """,
            (payload['user_id'], old_date, old_amount, old_category)
        )

        if len(rows) == 0:
            return {"status": "error", "message": "No matching expense found."}

        if len(rows) > 1:
            return {
                "status": "multiple_matches",
                "matches": [
                    {
                        "id": r['id'],
                        "date": r['date'],
                        "amount": r['amount'],
                        "category": r['category'],
                        "subcategory": r['subcategory'],
                        "note": r['note']
                    }
                    for r in rows
                ]
            }

        expense = rows[0]
        await conn.execute(
            """
            UPDATE expenses
            SET date=$1,
                amount=$2,
                category=$3,
                subcategory=$4,
                note=$5
            WHERE id=$6
            """,
            new_date if new_date is not None else expense['date'],
            new_amount if new_amount is not None else expense['amount'],
            new_category if new_category is not None else expense['category'],
            new_subcategory if new_subcategory is not None else expense['subcategory'],
            new_note if new_note is not None else expense['note'],
            expense['id']
        )
        return {"status": "ok", "expense_id": expense['id'], "message": "Expense updated successfully"}

    return await _run_with_connection(_op)


@mcp.tool(description="Record incoming money such as salary, reimbursement, cashback, refund, or bonus.")
async def add_credit(token: str, date: str, amount: float, source: str, note: str = ""):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            credit_id = await conn.fetchval(
                """
                INSERT INTO credits(user_id, date, amount, source, note)
                VALUES ($1, $2, $3, $4, $5)
                RETURNING id
                """,
                payload['user_id'], date, amount, source, note
            )
            return {"status": "ok", "credit_id": credit_id, "message": "Credit added successfully"}

        return await _run_with_connection(_op)


@mcp.tool(description="List all credited amounts within a date range.")
async def list_credits(token: str, start_date: str, end_date: str):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        rows = await _execute_fetchall(
            conn,
            """
            SELECT id, date, amount, source, note
            FROM credits
            WHERE user_id=$1 AND date BETWEEN $2 AND $3
            ORDER BY date ASC
            """,
            (payload['user_id'], start_date, end_date)
        )
        return [dict(row) for row in rows]

    return await _run_with_connection(_op)


@mcp.tool(description="Add a friend to the expense tracker so shared-expense emails can be sent to them.")
async def add_friend(token: str, name: str, email: str):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            await conn.execute("INSERT INTO friends(user_id,name,email) VALUES ($1,$2,$3)", payload['user_id'], name, email)
            return {"status": "ok", "message": f"{name} added"}

        return await _run_with_connection(_op)


@mcp.tool(description="Update the email address for an existing friend.")
async def update_friend_email(token: str, name: str, new_email: str):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            row = await _execute_fetchone(conn, "SELECT id FROM friends WHERE user_id=$1 AND LOWER(name)=LOWER($2)", (payload['user_id'], name))
            if not row:
                return {"status": "error", "message": f"Friend '{name}' not found"}
            await conn.execute("UPDATE friends SET email=$1 WHERE user_id=$2 AND LOWER(name)=LOWER($3)", new_email, payload['user_id'], name)
            return {"status": "ok", "message": f"Email updated for {name}"}

        return await _run_with_connection(_op)


@mcp.tool(description="List the friends that are currently registered in the tracker.")
async def list_friends(token: str):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        # Get current user info
        user_row = await _execute_fetchone(conn, "SELECT email, username FROM users WHERE id=$1", (payload['user_id'],))
        if not user_row:
            return []

        user_email, username = user_row['email'], user_row['username']

        # Get explicit friends (added by this user)
        explicit_rows = await _execute_fetchall(
            conn,
            "SELECT id, name, email FROM friends WHERE user_id=$1",
            (payload['user_id'],)
        )
        explicit_friends = [{'id': r['id'], 'name': r['name'], 'email': r['email']} for r in explicit_rows]

        # Get implicit friends (people who added this user as a friend)
        implicit_rows = await _execute_fetchall(
            conn, """
            SELECT DISTINCT f.user_id FROM friends f
            WHERE (
                (f.email IS NOT NULL AND LOWER(f.email) = LOWER($1))
                OR (f.name IS NOT NULL AND LOWER(f.name) = LOWER($2))
            )
            AND f.user_id != $3
            """,
            (user_email or '', username, payload['user_id'])
        )

        implicit_friends = []
        seen_emails = {f['email'] for f in explicit_friends if f['email']}

        for row in implicit_rows:
            adder_user_id = row['user_id']
            adder_row = await _execute_fetchone(
                conn,
                "SELECT username, email FROM users WHERE id=$1",
                (adder_user_id,)
            )
            if adder_row:
                adder_username, adder_email = adder_row['username'], adder_row['email']
                if adder_email and adder_email not in seen_emails:
                    implicit_friends.append({
                        'id': None,
                        'name': adder_username,
                        'email': adder_email
                    })
                    seen_emails.add(adder_email)

        return explicit_friends + implicit_friends

    return await _run_with_connection(_op)


@mcp.tool(description="Create a shared expense and split it among participants, including optional email notifications.")
async def add_shared_expense(token: str, date: str, amount: float, paid_by: str, participants: list, description: str = ""):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    participant_splits = await _build_participant_splits(amount, participants, paid_by)
    payer_name = await _normalize_payer_name(paid_by)

    async with db_lock:
        async def _op(conn):
            expense_id = await conn.fetchval(
                """
                INSERT INTO shared_expenses(user_id, date, description, total_amount, paid_by)
                VALUES ($1,$2,$3,$4,$5)
                RETURNING id
                """,
                payload['user_id'], date, description, amount, payer_name
            )

            for entry in participant_splits:
                await conn.execute(
                    """
                    INSERT INTO shared_expense_participants(expense_id, participant, share_amount)
                    VALUES ($1,$2,$3)
                    """,
                    expense_id, entry["name"], entry["share"]
                )

            balances = await _calculate_balances(conn, payload['user_id'])

            for entry in participant_splits:
                person = entry["name"]
                print(f"[EXPENSE] 🔍 Looking up email for participant: {person}")
                email_row = await _execute_fetchone(
                    conn, "SELECT email FROM friends WHERE user_id=$1 AND LOWER(name)=LOWER($2)",
                    (payload['user_id'], person)
                )

                if email_row and email_row['email']:
                    print(f"[EXPENSE] 📧 Found email: {email_row['email']}")
                    email_summary = await _build_email_summary(paid_by, amount, description, participant_splits, balances, person)
                    queue_email(email_row['email'], f"Expense Split: {description}", email_summary)
                else:
                    print(f"[EXPENSE] ⚠️  No email found for {person}")

            # Send email to the payer (current user) as well
            user_row = await _execute_fetchone(
                conn, "SELECT email FROM users WHERE id=$1",
                (payload['user_id'],)
            )
            if user_row and user_row['email']:
                print(f"[EXPENSE] 📧 Found payer email: {user_row['email']}")
                email_summary = await _build_email_summary(paid_by, amount, description, participant_splits, balances, "You")
                queue_email(user_row['email'], f"Expense Split: {description}", email_summary)
            else:
                print(f"[EXPENSE] ⚠️  No email found for payer")

            return {"status": "ok", "expense_id": expense_id, "splits": participant_splits}

        result = await _run_with_connection(_op)

    return result


@mcp.tool(description="List all shared expenses for the current user with participant details.")
async def list_shared_expenses(token: str, start_date: str = None, end_date: str = None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        if start_date and end_date:
            rows = await _execute_fetchall(
                conn,
                """
                SELECT id, date, description, total_amount, paid_by
                FROM shared_expenses
                WHERE user_id=$1 AND date BETWEEN $2 AND $3
                ORDER BY date DESC
                """,
                (payload['user_id'], start_date, end_date)
            )
        else:
            rows = await _execute_fetchall(
                conn,
                """
                SELECT id, date, description, total_amount, paid_by
                FROM shared_expenses
                WHERE user_id=$1
                ORDER BY date DESC
                """,
                (payload['user_id'],)
            )

        result = []
        for row in rows:
            expense_id = row['id']
            participants = await _execute_fetchall(
                conn,
                """
                SELECT participant, share_amount
                FROM shared_expense_participants
                WHERE expense_id=$1
                ORDER BY participant ASC
                """,
                (expense_id,)
            )

            result.append({
                "id": expense_id,
                "date": row['date'],
                "description": row['description'],
                "total_amount": row['total_amount'],
                "paid_by": row['paid_by'],
                "participants": [{"name": p['participant'], "share": p['share_amount']} for p in participants]
            })

        return result

    return await _run_with_connection(_op)


@mcp.tool(description="Record a settlement payment and notify the involved parties.")
async def settle_payment(token: str, person: str, amount: float, settlement_date: str, note: str = ""):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    email_tasks = []

    async with db_lock:
        async def _op(conn):
            await conn.execute(
                """
                INSERT INTO settlements(user_id, person, amount, settlement_date, note)
                VALUES ($1,$2,$3,$4,$5)
                """,
                payload['user_id'], person, amount, settlement_date, note
            )

            friend_row = await _execute_fetchone(conn, "SELECT email FROM friends WHERE user_id=$1 AND LOWER(name)=LOWER($2)", (payload['user_id'], person))
            if friend_row and friend_row['email']:
                queue_email(friend_row['email'], "Settlement recorded", f"Hi {person},\n\nA settlement of ₹{amount:.2f} was recorded on {settlement_date}.\n\nNote: {note or 'No note provided'}\n")

            return {"status": "ok", "message": "Settlement recorded"}

        result = await _run_with_connection(_op)

    return result


@mcp.tool(description="Calculate the net balance for each person from shared expenses and settlements.")
async def get_balances(token: str):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        return await _calculate_balances(conn, payload['user_id'])

    return await _run_with_connection(_op)


@mcp.tool(description="List all settlement records made by the current user.")
async def list_settlements(token: str):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        settlements = await _execute_fetchall(
            conn,
            "SELECT id, person, amount, settlement_date, note FROM settlements WHERE user_id=$1 ORDER BY settlement_date DESC",
            (payload['user_id'],)
        )
        return {
            "status": "ok",
            "settlements": [
                {
                    "id": s['id'],
                    "person": s['person'],
                    "amount": s['amount'],
                    "date": s['settlement_date'],
                    "note": s['note']
                }
                for s in settlements
            ]
        }

    return await _run_with_connection(_op)


@mcp.tool(description="Edit an existing expense (amount, category, date, note).")
async def edit_expense(token: str, expense_id: int, amount: float = None, category: str = None, date: str = None, note: str = None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            expense = await _execute_fetchone(
                conn, "SELECT id, user_id FROM expenses WHERE id=$1", (expense_id,)
            )
            if not expense or expense['user_id'] != payload['user_id']:
                return {"status": "error", "message": "Expense not found or unauthorized"}

            updates = []
            params = []
            param_count = 1

            if amount is not None:
                updates.append(f"amount=${param_count}")
                params.append(amount)
                param_count += 1
            if category is not None:
                updates.append(f"category=${param_count}")
                params.append(category)
                param_count += 1
            if date is not None:
                updates.append(f"date=${param_count}")
                params.append(date)
                param_count += 1
            if note is not None:
                updates.append(f"note=${param_count}")
                params.append(note)
                param_count += 1

            if not updates:
                return {"status": "error", "message": "No fields to update"}

            params.append(expense_id)
            query = f"UPDATE expenses SET {', '.join(updates)} WHERE id=${param_count} RETURNING id"

            result = await conn.fetchval(query, *params)
            return {"status": "ok", "expense_id": result, "message": "Expense updated successfully"}

        return await _run_with_connection(_op)


@mcp.tool(description="Delete an expense by ID.")
async def delete_expense(token: str, expense_id: int):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            expense = await _execute_fetchone(
                conn, "SELECT id, user_id, amount, category, date FROM expenses WHERE id=$1", (expense_id,)
            )
            if not expense or expense['user_id'] != payload['user_id']:
                return {"status": "error", "message": "Expense not found or unauthorized"}

            await conn.execute("DELETE FROM expenses WHERE id=$1", expense_id)
            return {
                "status": "ok",
                "message": "Expense deleted successfully",
                "deleted": {
                    "id": expense['id'],
                    "amount": expense['amount'],
                    "category": expense['category'],
                    "date": expense['date']
                }
            }

        return await _run_with_connection(_op)


@mcp.tool(description="Search and filter expenses by keyword, category, amount range, date range.")
async def search_expenses(token: str, keyword: str = None, category: str = None, min_amount: float = None, max_amount: float = None, start_date: str = None, end_date: str = None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        query = "SELECT id, date, amount, category, subcategory, note FROM expenses WHERE user_id=$1"
        params = [payload['user_id']]
        param_count = 2

        if keyword:
            query += f" AND (note ILIKE ${param_count} OR category ILIKE ${param_count} OR subcategory ILIKE ${param_count})"
            keyword_pattern = f"%{keyword}%"
            params.extend([keyword_pattern, keyword_pattern, keyword_pattern])
            param_count += 3

        if category:
            query += f" AND category=${param_count}"
            params.append(category)
            param_count += 1

        if min_amount is not None:
            query += f" AND amount>=${param_count}"
            params.append(min_amount)
            param_count += 1

        if max_amount is not None:
            query += f" AND amount<=${param_count}"
            params.append(max_amount)
            param_count += 1

        if start_date:
            query += f" AND date>=${param_count}"
            params.append(start_date)
            param_count += 1

        if end_date:
            query += f" AND date<=${param_count}"
            params.append(end_date)
            param_count += 1

        query += " ORDER BY date DESC"

        rows = await _execute_fetchall(conn, query, tuple(params))
        return {
            "status": "ok",
            "count": len(rows),
            "expenses": [dict(r) for r in rows]
        }

    return await _run_with_connection(_op)


@mcp.tool(description="Edit a shared expense (description, amount, date). When total_amount changes, participant shares are recalculated proportionally to maintain the same split ratio.")
async def edit_shared_expense(token: str, expense_id: int, description: str = None, total_amount: float = None, date: str = None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            expense = await _execute_fetchone(
                conn, "SELECT id, user_id, total_amount FROM shared_expenses WHERE id=$1", (expense_id,)
            )
            if not expense or expense['user_id'] != payload['user_id']:
                return {"status": "error", "message": "Shared expense not found or unauthorized"}

            original_total = float(expense['total_amount'])

            updates = []
            params = []
            param_count = 1

            if description is not None:
                updates.append(f"description=${param_count}")
                params.append(description)
                param_count += 1
            if total_amount is not None:
                updates.append(f"total_amount=${param_count}")
                params.append(total_amount)
                param_count += 1
            if date is not None:
                updates.append(f"date=${param_count}")
                params.append(date)
                param_count += 1

            if not updates:
                return {"status": "error", "message": "No fields to update"}

            params.append(expense_id)
            query = f"UPDATE shared_expenses SET {', '.join(updates)} WHERE id=${param_count} RETURNING id, total_amount"

            result = await conn.fetchrow(query, *params)
            new_total = float(result['total_amount']) if result else original_total

            # If total_amount was changed, recalculate participant shares proportionally
            if total_amount is not None and original_total > 0:
                participants = await _execute_fetchall(
                    conn,
                    "SELECT id, participant, share_amount FROM shared_expense_participants WHERE expense_id=$1",
                    (expense_id,)
                )

                for participant in participants:
                    original_share = float(participant['share_amount'])
                    proportion = original_share / original_total
                    new_share = round(new_total * proportion, 2)

                    await conn.execute(
                        "UPDATE shared_expense_participants SET share_amount=$1 WHERE id=$2",
                        new_share,
                        participant['id']
                    )

            return {
                "status": "ok",
                "expense_id": result['id'] if result else expense_id,
                "message": "Shared expense updated successfully",
                "shares_recalculated": total_amount is not None
            }

        return await _run_with_connection(_op)


@mcp.tool(description="Delete a shared expense by ID.")
async def delete_shared_expense(token: str, expense_id: int):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async with db_lock:
        async def _op(conn):
            expense = await _execute_fetchone(
                conn, "SELECT id, user_id, description, total_amount, date FROM shared_expenses WHERE id=$1", (expense_id,)
            )
            if not expense or expense['user_id'] != payload['user_id']:
                return {"status": "error", "message": "Shared expense not found or unauthorized"}

            await conn.execute("DELETE FROM shared_expense_participants WHERE expense_id=$1", expense_id)
            await conn.execute("DELETE FROM shared_expenses WHERE id=$1", expense_id)

            return {
                "status": "ok",
                "message": "Shared expense deleted successfully",
                "deleted": {
                    "id": expense['id'],
                    "description": expense['description'],
                    "total_amount": expense['total_amount'],
                    "date": expense['date']
                }
            }

        return await _run_with_connection(_op)


@mcp.tool(description="Search and filter shared expenses by keyword, amount range, date range, paid_by.")
async def search_shared_expenses(token: str, keyword: str = None, paid_by: str = None, min_amount: float = None, max_amount: float = None, start_date: str = None, end_date: str = None):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    async def _op(conn):
        query = "SELECT id, date, description, total_amount, paid_by FROM shared_expenses WHERE user_id=$1"
        params = [payload['user_id']]
        param_count = 2

        if keyword:
            query += f" AND (description ILIKE ${param_count})"
            keyword_pattern = f"%{keyword}%"
            params.append(keyword_pattern)
            param_count += 1

        if paid_by:
            query += f" AND LOWER(paid_by)=LOWER(${param_count})"
            params.append(paid_by)
            param_count += 1

        if min_amount is not None:
            query += f" AND total_amount>=${param_count}"
            params.append(min_amount)
            param_count += 1

        if max_amount is not None:
            query += f" AND total_amount<=${param_count}"
            params.append(max_amount)
            param_count += 1

        if start_date:
            query += f" AND date>=${param_count}"
            params.append(start_date)
            param_count += 1

        if end_date:
            query += f" AND date<=${param_count}"
            params.append(end_date)
            param_count += 1

        query += " ORDER BY date DESC"

        rows = await _execute_fetchall(conn, query, tuple(params))

        result_expenses = []
        for row in rows:
            participants = await _execute_fetchall(
                conn,
                "SELECT participant, share_amount FROM shared_expense_participants WHERE expense_id=$1",
                (row['id'],)
            )
            result_expenses.append({
                "id": row['id'],
                "date": row['date'],
                "description": row['description'],
                "total_amount": row['total_amount'],
                "paid_by": row['paid_by'],
                "participants": [{"name": p['participant'], "share": p['share_amount']} for p in participants]
            })

        return {
            "status": "ok",
            "count": len(result_expenses),
            "shared_expenses": result_expenses
        }

    return await _run_with_connection(_op)


@mcp.tool(description="Export expenses to Excel file. Filter by category, date range, type (personal/shared).")
async def export_expenses_to_excel(token: str, start_date: str = None, end_date: str = None, category: str = None, include_shared: bool = True):
    payload = _verify_token(token)
    if not payload:
        return {"status": "error", "message": "Invalid or expired token"}

    if Workbook is None:
        return {"status": "error", "message": "Excel library not installed. Run: pip install openpyxl"}

    async def _op(conn):
        user = await _execute_fetchone(conn, "SELECT username FROM users WHERE id=$1", (payload['user_id'],))
        username = user['username'] if user else "User"

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["ID", "Date", "Amount", "Category", "Subcategory", "Note", "Type"]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        query = "SELECT id, date, amount, category, subcategory, note FROM expenses WHERE user_id=$1"
        params = [payload['user_id']]
        param_count = 2

        if start_date:
            query += f" AND date>=${param_count}"
            params.append(start_date)
            param_count += 1

        if end_date:
            query += f" AND date<=${param_count}"
            params.append(end_date)
            param_count += 1

        if category:
            query += f" AND category=${param_count}"
            params.append(category)
            param_count += 1

        query += " ORDER BY date DESC"
        expenses = await _execute_fetchall(conn, query, tuple(params))

        for expense in expenses:
            ws.append([
                expense['id'],
                expense['date'],
                expense['amount'],
                expense['category'],
                expense['subcategory'],
                expense['note'],
                "Personal"
            ])

        if include_shared:
            ws_shared = wb.create_sheet("Shared Expenses")
            ws_shared.append(headers + ["Paid By", "Participants"])

            for cell in ws_shared[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            shared_query = """
                SELECT se.id, se.date, se.total_amount, se.description, se.paid_by
                FROM shared_expenses se
                WHERE se.user_id=$1
            """
            shared_params = [payload['user_id']]
            if start_date:
                shared_query += f" AND se.date>={len(shared_params)+1}"
                shared_params.append(start_date)
            if end_date:
                shared_query += f" AND se.date<={len(shared_params)+1}"
                shared_params.append(end_date)
            shared_query += " ORDER BY se.date DESC"

            shared_expenses = await _execute_fetchall(conn, shared_query, tuple(shared_params))

            for se in shared_expenses:
                participants = await _execute_fetchall(
                    conn,
                    "SELECT participant FROM shared_expense_participants WHERE expense_id=$1",
                    (se['id'],)
                )
                participant_names = ", ".join([p['participant'] for p in participants])
                ws_shared.append([
                    se['id'],
                    se['date'],
                    se['total_amount'],
                    se['description'],
                    "",
                    "",
                    "Shared",
                    se['paid_by'],
                    participant_names
                ])

        for ws in wb.sheetnames:
            for column in wb[ws].columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                wb[ws].column_dimensions[column_letter].width = adjusted_width

        file_path = f"/tmp/expenses_{username}_{datetime.now(tz.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(file_path)

        return {
            "status": "ok",
            "message": "Excel file exported successfully",
            "file_path": file_path,
            "file_size": os.path.getsize(file_path),
            "sheets": ["Expenses", "Shared Expenses"] if include_shared else ["Expenses"]
        }

    return await _run_with_connection(_op)


@mcp.resource("expense://categories", mime_type="application/json")
async def categories():
    categories_file = os.path.join(os.getcwd(), "categories.json")
    if not os.path.exists(categories_file):
        return "[]"
    async with aiofiles.open(categories_file, "r", encoding="utf-8") as f:
        return await f.read()


app = Flask(__name__)
CORS(app, resources={r"/mcp/*": {"origins": "*"}})
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def get_tool_schema(tool):
    """Extract schema from a FastMCP tool"""
    if hasattr(tool, '__doc__'):
        doc = tool.__doc__ or ""
    else:
        doc = ""

    import inspect
    sig = inspect.signature(tool)
    properties = {}
    required = []

    for param_name, param in sig.parameters.items():
        if param_name == 'self':
            continue
        properties[param_name] = {
            "type": "string",
            "description": f"Parameter {param_name}"
        }
        if param.default == inspect.Parameter.empty:
            required.append(param_name)

    return {
        "type": "object",
        "properties": properties,
        "required": required
    }


@app.route('/mcp/tools', methods=['GET'])
def get_tools():
    """Return available tools from MCP server"""
    tools = []
    for tool_name, tool_func in _tool_registry.items():
        tools.append({
            "name": tool_name,
            "description": (tool_func.__doc__ or "").strip(),
            "inputSchema": get_tool_schema(tool_func)
        })
    return jsonify({"tools": tools})


@app.route('/mcp/call_tool', methods=['POST'])
def call_tool():
    """Execute a tool from the MCP server"""
    data = request.json
    tool_name = data.get('name')
    arguments = data.get('arguments', {})

    if not tool_name:
        return jsonify({"error": "Tool name required"}), 400

    if tool_name not in _tool_registry:
        logger.warning(f"Unknown tool requested: {tool_name}")
        logger.info(f"Available tools: {list(_tool_registry.keys())}")
        return jsonify({"error": f"Unknown tool: {tool_name}"}), 404

    try:
        tool_func = _tool_registry[tool_name]
        # Check if the tool is async
        if asyncio.iscoroutinefunction(tool_func):
            result = asyncio.run(tool_func(**arguments))
        else:
            result = tool_func(**arguments)
        return jsonify({"result": result, "success": True})
    except Exception as e:
        logger.error(f"Tool execution error for {tool_name}: {e}", exc_info=True)
        return jsonify({"error": str(e), "success": False}), 500


@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({"status": "ok"})


@app.route('/', methods=['GET'])
def index():
    """Root endpoint"""
    return jsonify({"message": "Expense Tracker MCP Server", "status": "running"})


if __name__ == "__main__":
    port = int(os.environ.get('PORT', 8000))
    app.run(host='0.0.0.0', port=port, debug=False)
